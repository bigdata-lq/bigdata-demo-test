import pandas as pd
from openpyxl import load_workbook
from etl.testing.lqTest.tools.read_config import ReadConfig
from etl.testing.lqTest.tools.do_regx import DoRegx
from etl.testing.lqTest.tools.my_log import MyLog
from collections import namedtuple
from pprint import pprint
from jsonpath import jsonpath
import json
import ast
from etl.testing.lqTest.tools.do_database import run_mysql
from faker import Faker

f = Faker(locale='zh_CN')
my_logger = MyLog()

#openpyxl操作excel时，行号和列号都是从1开始计算的
class DoExcel:

    def __init__(self,filename,sheet_name):
        try:
            self.filename = filename
            self.sheet_name = sheet_name
            self.wb = load_workbook(self.filename)
            if self.sheet_name is None:
                self.work_sheet = self.wb.active
            else:
                self.work_sheet = self.wb[self.sheet_name]
        except FileNotFoundError as e:
            my_logger.error("文件不存在")
            raise e


    def get_max_row_num(self):
        """获取最大行号"""
        max_row_num = self.work_sheet.max_row
        return max_row_num

    def get_max_column_num(self):
        """获取最大列号"""
        max_column = self.work_sheet.max_column
        return max_column

    def get_cell_value(self, coordinate=None, row=None, column=None):
        """获取指定单元格的数据    coordinate为单元格的位置，例如A1  B3 等"""
        if coordinate is not None:
            try:
                return self.work_sheet[coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and row is not None and column is not None:
            if isinstance(row, int) and isinstance(column, int):
                return self.work_sheet.cell(row=row, column=column).value
            else:
                raise TypeError('row and column must be type int')
        else:
            raise Exception("Insufficient Coordinate of cell!")

    def get_row_value(self, row:int):
        """获取某一行的数据"""
        column_num = self.get_max_column_num()
        row_value = []
        if isinstance(row, int):
            for column in range(1, column_num + 1):
                values_row = self.work_sheet.cell(row, column).value
                row_value.append(values_row)
            return row_value
        else:
            raise TypeError('row must be type int')

    def get_column_value(self, column:int):
        """获取某一列数据"""
        row_num = self.get_max_column_num()
        # print(row_num)
        column_value = []
        if isinstance(column, int):
            for row in range(1, row_num + 1):
                values_column = self.work_sheet.cell(row, column).value
                column_value.append(values_column)
            return column_value
        else:
            raise TypeError('column must be type int')



    def get_column_value_xk(self,column:str):
        """
            获取某一列数据   列头，如，A/B/C   小写的也行
        :param column:  列头，如，A/B/C   小写的也行
        :return:
        """
        # wb=load_workbook(r"E:\Program Files (x86)\download\APItest_unittest_cmp-master\test_data\case.xlsx")
        # ws = wb.get_sheet_by_name('init')
        first_column = self.work_sheet[column]
        # print(len(first_column))
        column_value = []
        for x in range(len(first_column)):
            # if first_column[x].value != None:
                column_value.append(first_column[x].value)
        return column_value


    def get_all_value_1(self):
        """获取指定表单的所有数据(除去表头)"""
        max_row_num = self.get_max_row_num()
        max_column = self.get_max_column_num()
        values = []
        for row in range(2, max_row_num + 1):
            value_list = []
            for column in range(1, max_column + 1):
                value = self.work_sheet.cell(row, column).value
                value_list.append(value)
            values.append(value_list)
        return values


    def get_all_value_2(self):
        """获取指定表单的所有数据(除去表头)"""
        rows_obj = self.work_sheet.iter_rows(min_row=2, max_row=self.work_sheet.max_row,
                                             values_only=True)  # 指定values_only 会直接提取数据不需要再使用cell().value
        # print(rows_obj)
        values = []
        for row_tuple in rows_obj:
            value_list = []
            for value in row_tuple:
                value_list.append(value)
            values.append(value_list)
        return values


    def get_excel_title(self):
        """获取sheet表头"""
        title_key = tuple(self.work_sheet.iter_rows(max_row=1, values_only=True))[0]
        return title_key


    def get_listdict_all_value(self):
        """获取所有数据，返回嵌套字典的列表"""
        sheet_title = self.get_excel_title()
        all_values = self.get_all_value_2()
        value_list = []
        for value in all_values:
            value_list.append(dict(zip(sheet_title, value)))
            # print(DoRegx.do_regx(str(dict(zip(sheet_title, value)))))
        return value_list
    #
    # def get_list_nametuple_all_value(self):
    #     """获取所有数据，返回嵌套命名元组的列表"""
    #     sheet_title = self.get_excel_title()
    #     values = self.get_all_value_2()
    #
    #     excel = namedtuple('excel', sheet_title)
    #     value_list = []
    #     for value in values:
    #         e = excel(*value)
    #         value_list.append(e)
    #     return value_list


    #获取sheet页所有数据
    # def get_data(self,filename,sheetname):

    def get_all_data(self):
        wb =load_workbook(self.filename)
        # mode = eval(ReadConfig.get_config(case_config_path,"MODE","mode"))
        sheet = wb[self.sheet_name]  # 表单名
        test_data =[]
        for case_id in range(2, sheet.max_row+1):
            # print(case_id,"行")
            row_data = {}  # 字典
            is_run = sheet.cell(None, case_id, int(ReadConfig.get_test_data("run"))).value
            # print(type(is_run))
            #通过 excel里的run列判断是否把数据加进DDT去执行
            if is_run == 'yes':
                row_data["case_id"] = sheet.cell(None,case_id, int(ReadConfig.get_test_data("case_id"))).value
                row_data["module"] = sheet.cell(None,case_id,int(ReadConfig.get_test_data("module"))).value
                row_data["title"] = sheet.cell(None,case_id, int(ReadConfig.get_test_data("title"))).value
                # row_data["url"] = sheet.cell(None,case_id, int(ReadConfig.get_test_data("url"))).value
                #正则替换
                row_data["url"] = DoRegx.do_regx(sheet.cell(None,case_id, int(ReadConfig.get_test_data("url"))).value)

                # row_data["data"] = sheet.cell(case_id,int(ReadConfig.get_test_data("request_data"))).value
                #非正则字串符替换
                #替换管理员账户
                # if sheet.cell(case_id+1, int(ReadConfig.get_test_data("request_data"))).value.find("${user}") != -1 and sheet.cell(case_id, int(ReadConfig.get_test_data("request_data"))).value.find("${passwd}") != -1:  # if h后面非空  非零  成立的表达式  都为True，只要是True，if下面的代码都会执行
                #     row_data["data"] = sheet.cell(case_id, int(ReadConfig.get_test_data("request_data"))).value.replace("${user}", getattr(GetData,"admin_user")).replace("${passwd}", str(getattr(GetData,"admin_passwd")))
                # else:
                #     row_data["data"] = sheet.cell(case_id+1, int(ReadConfig.get_test_data("request_data"))).value

                #正则字符串替换   一行替代多行
                row_data["data"] = DoRegx.do_regx(sheet.cell(None,case_id, int(ReadConfig.get_test_data("request_data"))).value)
                row_data["run"] = sheet.cell(None,case_id, int(ReadConfig.get_test_data("run"))).value
                row_data["http_method"] = sheet.cell(None,case_id,int(ReadConfig.get_test_data("http_method"))).value

                # sql 语句中的字符串替换
                # if sheet.cell(case_id+1, int(ReadConfig.get_test_data("check_sql"))).value !=None:
                #     if sheet.cell(case_id+1,int(ReadConfig.get_test_data("check_sql"))).value.find("${normal_tel}") != -1:
                #         row_data["data"] = sheet.cell(case_id+1, int(ReadConfig.get_test_data("request_data"))).value.replace("${user}", getattr(GetData, "admin_user"))
                # else:
                #     row_data["data"] = DoRegx.do_regx(str(sheet.cell(case_id+1, int(ReadConfig.get_test_data("request_data"))).value))

                row_data["check_sql"] = DoRegx.do_regx(str(sheet.cell(None,case_id, int(ReadConfig.get_test_data("check_sql"))).value))
                row_data["expect"] = sheet.cell(None,case_id, int(ReadConfig.get_test_data("expect"))).value
                row_data["sheet_name"] = self.sheet_name
                row_data["json_path"] = sheet.cell(None,case_id, int(ReadConfig.get_test_data("json_path"))).value
                row_data["headers"] = DoRegx.do_regx(sheet.cell(None,case_id, int(ReadConfig.get_test_data("headers"))).value)
                test_data.append(row_data)

            # else:
            #     row_data = {}  # 字典
            #     row_data["case_id"] = sheet.cell(case_id+1, 1).value
            #     row_data["username"] = sheet.cell(case_id+1, 2).value
            #     row_data["password"] = sheet.cell(case_id+1, 3).value
            #     row_data["email"] = sheet.cell(i, 4).value
            #     row_data["sheet_name"] = sheetname
            #     test_data.append(row_data)

        return test_data


    def get_row_data_xk(self,row:int):
        wb =load_workbook(self.filename)
        # mode = eval(ReadConfig.get_config(case_config_path,"MODE","mode"))
        sheet = wb[self.sheet_name]  # 表单名

        list_row_data = []
        row_data = {}  # 字典
        is_run = sheet.cell(row, int(ReadConfig.get_test_data("run"))).value
        # print(type(is_run))
        #通过 excel里的run列判断是否把数据加进DDT去执行
        if is_run == 'yes':
            row_data["case_id"] = sheet.cell(None,row, int(ReadConfig.get_test_data("case_id"))).value
            row_data["module"] = sheet.cell(None,row,int( ReadConfig.get_test_data("module"))).value
            row_data["title"] = sheet.cell(None,row, int(ReadConfig.get_test_data("title"))).value
            # row_data["url"] = sheet.cell(None,case_id, int(ReadConfig.get_test_data("url"))).value
            #正则替换
            row_data["url"] = DoRegx.do_regx(sheet.cell(row, int(ReadConfig.get_test_data("url"))).value)

            # row_data["data"] = sheet.cell(case_id,int(ReadConfig.get_test_data("request_data"))).value
            #非正则字串符替换
            #替换管理员账户
            # if sheet.cell(case_id+1, int(ReadConfig.get_test_data("request_data"))).value.find("${user}") != -1 and sheet.cell(case_id, int(ReadConfig.get_test_data("request_data"))).value.find("${passwd}") != -1:  # if h后面非空  非零  成立的表达式  都为True，只要是True，if下面的代码都会执行
            #     row_data["data"] = sheet.cell(case_id, int(ReadConfig.get_test_data("request_data"))).value.replace("${user}", getattr(GetData,"admin_user")).replace("${passwd}", str(getattr(GetData,"admin_passwd")))
            # else:
            #     row_data["data"] = sheet.cell(case_id+1, int(ReadConfig.get_test_data("request_data"))).value

            #正则字符串替换   一行替代多行
            row_data["data"] = DoRegx.do_regx(sheet.cell(None,row, int(ReadConfig.get_test_data("request_data"))).value)
            # row_data["data"] = sheet.cell(row, int(ReadConfig.get_test_data("request_data"))).value
            row_data["run"] = sheet.cell(None,row, int(ReadConfig.get_test_data("run"))).value
            row_data["http_method"] = sheet.cell(None,row,int(ReadConfig.get_test_data("http_method"))).value

            # sql 语句中的字符串替换
            # if sheet.cell(case_id+1, int(ReadConfig.get_test_data("check_sql"))).value !=None:
            #     if sheet.cell(case_id+1,int(ReadConfig.get_test_data("check_sql"))).value.find("${normal_tel}") != -1:
            #         row_data["data"] = sheet.cell(case_id+1, int(ReadConfig.get_test_data("request_data"))).value.replace("${user}", getattr(GetData, "admin_user"))
            # else:
            #     row_data["data"] = DoRegx.do_regx(str(sheet.cell(case_id+1, int(ReadConfig.get_test_data("request_data"))).value))

            row_data["check_sql"] = DoRegx.do_regx(str(sheet.cell(None,row, int(ReadConfig.get_test_data("check_sql"))).value))
            row_data["expect"] = sheet.cell(None,row, int(ReadConfig.get_test_data("expect"))).value
            row_data["sheet_name"] = self.sheet_name
            row_data["json_path"] = sheet.cell(None,row, int(ReadConfig.get_test_data("json_path"))).value
            row_data["headers"] = DoRegx.do_regx(sheet.cell(None,row, int(ReadConfig.get_test_data("headers"))).value)
        list_row_data.append(row_data)
        return list_row_data
        # else:
        #     row_data = {}  # 字典
        #     row_data["case_id"] = sheet.cell(case_id+1, 1).value
        #     row_data["username"] = sheet.cell(case_id+1, 2).value
        #     row_data["password"] = sheet.cell(case_id+1, 3).value
        #     row_data["email"] = sheet.cell(i, 4).value
        #     row_data["sheet_name"] = sheetname
        #     test_data.append(row_data)

        # return row_data


    def get_json_path_value(self,res,variable_name = "userToken",json_path = '$.data.userToken'):
        """
            类似于 jmeter 中的json提取器,并写入 init 文件，如果变量名存在则替换原值，不存在则新增在最后一行+1
        :param res:              接口响应结果
        :param variable_name:    变量名
        :param json_path:        json_path
        :return:
        """
        json_path_value = jsonpath(res.json(),json_path)
        if json_path_value != False:
            json_path_value = json_path_value[0]
        else:
            print("获取json_path_value失败")
            return
        if variable_name  in tuple(DoExcel(test_case_data_path,"init").get_column_value_xk("A")):
            index = DoExcel(test_case_data_path,"init").get_column_value_xk("A").index(variable_name)
            DoExcel(test_case_data_path,"init").write_back(index + 1,2,json_path_value)
        else:
            max_row_num = self.get_max_row_num()
            DoExcel(test_case_data_path,"init").write_back(max_row_num + 1,1,variable_name)
            DoExcel(test_case_data_path,"init").write_back(max_row_num + 1,2,json_path_value)





    def get_json_path_value_1(self,res,list_json_path = None):         # 这个有空debug看看问题出在哪，怎么就写入一个变量呢，运行第二次就写全了。。。
        """
            类似于 jmeter 中的json提取器,并写入 init 文件，如果变量名存在则替换原值，不存在则新增在最后一行+1
        :param res:              接口响应结果
        :param list_json_path:    变量名&json_path
        :return:
        """
        # list_json_path = ast.literal_eval(DoExcel(test_case_data_path,"login").get_cell_value(coordinate="M2"))
        if list_json_path != None:
            max_row_num = self.get_max_row_num()
            for i in range(len(list_json_path)):
                max_row_num += 1
                for key in list_json_path[i]:
                    # print(key,list_json_path[i][key])
                    json_path_value = jsonpath(res.json(),list_json_path[i][key])
                    if json_path_value != False:
                        # print("json_path_value:",json_path_value)
                        new_json_path_value = json_path_value[0]
                        if key in tuple(self.get_column_value_xk("A")):
                            index = self.get_column_value_xk("A").index(key)
                            self.write_back(index + 1,2,new_json_path_value)
                        else:
                            self.write_back(max_row_num,1,key)
                            self.write_back(max_row_num,2,new_json_path_value)
                    else:
                        return
        else:
            pass


    def write_init_data_by_faker(self,list_keys_values:list = None):         # 这个有空debug看看问题出在哪，怎么就写入一个变量呢，运行第二次就写全了。。。
        """
            随机数据--->>>注意：有局限性
        :param list_keys_values:
        :return:
        """
        # list_json_path = ast.literal_eval(DoExcel(test_case_data_path,"login").get_cell_value(coordinate="M2"))
        if list_keys_values != None:
            max_row_num = self.get_max_row_num()
            for i in range(len(list_keys_values)):
                max_row_num += 1
                for key in list_keys_values[i]:
                    # print(key,list_keys_values[i][key])
                    if key in tuple(self.get_column_value_xk("A")):
                        index = self.get_column_value_xk("A").index(key)
                        self.write_back(index + 1,2,list_keys_values[i][key])
                    else:
                        self.write_back(max_row_num,1,key)
                        self.write_back(max_row_num,2,list_keys_values[i][key])

        else:
            pass


    def run_mysql(self,list_sql = None):
        # 有的sql只是查询结果，有的sql是查询到结果还要写入sheet当变量用----->>> 怎么实现呢!!!!!
        if list_sql != None:
            pass

        else:
            pass



    # def write_back(filename,sheetname,row,col,result):#回写数据
    def write_back(self,row:int, col:int, result):
        try:
            if isinstance(row,int) and isinstance(col,int):
                wb = load_workbook(self.filename)
                sheet = wb[self.sheet_name]
                sheet.cell(None,row,col).value = result
                wb.save(self.filename)#保存结果
                # wb.close()
            else:
                raise TypeError("row and col 必须是 int类型")
        except Exception as e:
            raise e








if __name__ == '__main__':
    from dataconfig.project_path import *

    print(DoExcel(test_case_data_path,"login").get_all_data())
    # print(DoExcel(test_case_data_path,"login").get_excel_title())
    # print(DoExcel(test_case_data_path,"login").get_all_value_2())
    # print(DoExcel(test_case_data_path,"login").get_cell_value(row=2,column=2))


    # print(DoExcel(test_case_data_path,"login").get_cell_value(coordinate="E15"))

    # print(DoExcel(test_case_data_path,"login").get_row_value(2))
    # print(DoExcel(test_case_data_path,"init").get_max_column_num())
    # print(DoExcel(test_case_data_path,"init").get_max_row_num())

    # print(DoExcel(test_case_data_path,"init").get_column_value(1))

    # from openpyxl import load_workbook
    # wb=load_workbook(r"E:\Program Files (x86)\download\APItest_unittest_cmp-master\test_data\case.xlsx")
    # ws = wb.get_sheet_by_name('init')
    # first_column = ws['a']
    # print(len(first_column))
    # for x in range(len(first_column)):
    #     if first_column[x].value != None:
    #         print(first_column[x].value)


    # print(DoExcel(test_case_data_path,"init").get_column_value_xk("A"))







    # if "virtual_sales"  in tuple(DoExcel(test_case_data_path,"init").get_column_value_xk("A")):
    #             index = DoExcel(test_case_data_path,"init").get_column_value_xk("A").index("virtual_sales")
    #             DoExcel(test_case_data_path,"init").write_back(index + 1,2,"pp我爱你")




    import requests
    res =  requests.post(
              url="https://union-uat.zgmmtuan.com/admin/Index/login",
              data={
                  "account":'ceshi1',
                  "password":123
              },
              headers = {
                  # "userToken":"d624e52d2cabfa5f68a90c7912014723",
                  "Content-Type":"application/x-www-form-urlencoded"
              },
              verify = False
    )
    # print(res.json())
    # print(type(json.dumps(res.json())))
    #
    # res1 =  requests.post(
    #           url="https://union-uat.zgmmtuan.com/admin/Capital/rerge",
    #           data={
    #               "user_id":77,
    #               "money":10,
    #               "type":0
    #           },
    #           headers = {
    #               "token":"{0}".format(res.json()["data"]["user_token"]),
    #               "Content-Type":"application/x-www-form-urlencoded"
    #           },
    #           verify = False
    # )
    # print(res1.json())
    # print("====="*400)
    # # print(DoExcel(test_case_data_path,"login").get_row_data_xk(3))
    # item = DoExcel(test_case_data_path,"login").get_row_data_xk(2)
    # print(item)


    # DoExcel(test_case_data_path,"init").write_back(54,2, str(res.json()))

    # print(type(ast.literal_eval(DoExcel(test_case_data_path,"init").get_cell_value(coordinate="B54"))))
    # print(ast.literal_eval(DoExcel(test_case_data_path,"init").get_cell_value(coordinate="B54")))

    # DoExcel(test_case_data_path,"init").get_json_path_value(res,"xpp","$.msg")

    # print(ast.literal_eval(DoExcel(test_case_data_path,"login").get_cell_value(coordinate="M2")))
    #
    DoExcel(test_case_data_path,"init").get_json_path_value_1(res,list_json_path = [{'houtai_token': '$.data.user_token'}])
    DoExcel(test_case_data_path,"init").write_init_data_by_faker(list_keys_values=[{"key001":f.name()},{"key002":f.name()}])
"""
    # 字符串列表转换成列表
    print(ast.literal_eval(DoExcel(test_case_data_path,"login").get_cell_value(coordinate="M2")))
    print(ast.literal_eval(DoExcel(test_case_data_path,"login").get_cell_value(coordinate="M2"))[0]["coupon_id_1"])
    print(type(ast.literal_eval(DoExcel(test_case_data_path,"login").get_cell_value(coordinate="M2"))[0]))
"""
